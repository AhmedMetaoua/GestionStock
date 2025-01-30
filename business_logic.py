from db_manager import get_connection
import json
from datetime import datetime
import openpyxl
import xlwings as xw
from num2words import num2words
import os
from pathlib import Path

def ajouter_matiere_premiere(nom, reference, quantite, prix):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM matieres_premieres WHERE nom = ? and reference = ?", (nom, reference,))
    matiere_premiere_id = cursor.fetchone()
    if matiere_premiere_id:
        conn.close()
        raise ValueError(f"Matière première '{nom}' existe déja.")
    cursor.execute("INSERT INTO matieres_premieres (nom, reference, quantite, prix_unitaire) VALUES (?, ?, ?, ?)", (nom, reference, quantite, prix))
    conn.commit()
    conn.close()

def modifier_quantite_matiere_premiere(identifiant, quantite):
    conn = get_connection()
    cursor = conn.cursor()
    
    #  Mettre à jour par nom
    cursor.execute("UPDATE matieres_premieres SET quantite = quantite + ? WHERE nom = ?", (quantite, identifiant))
    if cursor.rowcount == 0:
        conn.close()
        raise ValueError(f"Matière première avec identifiant '{identifiant}' introuvable.")

    conn.commit()
    conn.close()

def mettre_aJour_produit_fini(identifiant, quantite, prix, fodec_value):
    conn = get_connection()
    cursor = conn.cursor()

    # verfier si le produit existe on va faire un update sinon on fait une insertion
    cursor.execute("SELECT reference FROM matieres_produites WHERE nom = ?", (identifiant,))
    result = cursor.fetchone()
    matiere_produite_reference = result[0] if result else None

    if not matiere_produite_reference:
        cursor.execute("INSERT INTO matieres_produites (nom, reference, quantite, prix_unitaire, fodec) VALUES (?, ?, ?, ?, ?)",
                    (identifiant, f"REF-{identifiant.upper()}", quantite, prix, fodec_value))  
    else :
        cursor.execute("UPDATE matieres_produites SET quantite = quantite + ?, prix_unitaire = ?, fodec = ? WHERE nom = ?", 
                       (quantite, prix, fodec_value, identifiant))

    conn.commit()
    conn.close()

def ajouter_dosage(nom_matiere_produite, dosages):
    """
    Ajoute des dosages pour une matière produite avec plusieurs matières premières.

    :param nom_matiere_produite: Nom de la matière produite
    :param dosages: Liste de tuples (nom_matiere_premiere, proportion)
    """
    conn = get_connection()
    cursor = conn.cursor()

    # Référence de la matière produite
    matiere_produite_reference = f"REF-{nom_matiere_produite.upper()}"

    for nom_matiere_premiere, proportion in dosages:
        # Récupérer l'ID de la matière première
        cursor.execute("SELECT id FROM matieres_premieres WHERE nom = ?", (nom_matiere_premiere,))
        matiere_premiere_id = cursor.fetchone()
        if not matiere_premiere_id:
            conn.close()
            raise ValueError(f"Matière première '{nom_matiere_premiere}' introuvable.")
        matiere_premiere_id = matiere_premiere_id[0]

        # Vérifier si ce dosage existe déjà
        cursor.execute("""SELECT id FROM dosages WHERE matiere_produite_reference = ? AND matiere_premiere_id = ?""", 
                       (matiere_produite_reference, matiere_premiere_id))
        if cursor.fetchone():
            cursor.execute("UPDATE dosages SET proportion = ? WHERE matiere_produite_reference = ? and matiere_premiere_id = ?", 
                           (proportion, matiere_produite_reference, matiere_premiere_id))
        else:
            # Ajouter le dosage
            cursor.execute("INSERT INTO dosages (matiere_produite_reference, matiere_premiere_id, proportion) VALUES (?, ?, ?)",
                           (matiere_produite_reference, matiere_premiere_id, proportion))

    conn.commit()
    conn.close()

def ajouter_client(nom, adresse, phone, reference):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM clients WHERE nom = ? AND reference = ?", (nom, reference,))
    client_existant = cursor.fetchone()
    if client_existant:
        conn.close()
        raise ValueError(f"Le client '{nom}' existe déjà avec cette référence.")
    cursor.execute("INSERT INTO clients (nom, adresse, phone, reference) VALUES (?, ?, ?, ?)", (nom, adresse, phone, reference))
    conn.commit()
    conn.close()

def creer_bon_de_commande(nom, quantite, prix, fodec_value):
    conn = get_connection()
    cursor = conn.cursor()

    # Récupérer les dosages pour cette matière produite
    cursor.execute("SELECT reference FROM matieres_produites WHERE nom = ?", (nom,))
    result = cursor.fetchone()
    matiere_produite_reference = result[0] if result else None

    if not matiere_produite_reference:
        cursor.execute("INSERT INTO matieres_produites (nom, reference, quantite, prix_unitaire, fodec) VALUES (?, ?, ?, ?, ?)",
                    (nom, f"REF-{nom.upper()}", 0, prix, fodec_value))  # Initialise avec quantité = 0
        matiere_produite_reference = f"REF-{nom.upper()}"

    cursor.execute("SELECT matiere_premiere_id, proportion FROM dosages WHERE matiere_produite_reference = ?", (matiere_produite_reference,))
    dosages = cursor.fetchall()
    if not dosages:
        raise ValueError(f"Erreur : Aucun dosage trouvé pour la matière produite : {nom}")

    # Liste pour garder trace des modifications déjà effectuées
    modifications = []

    # Vérifier et réduire les quantités
    for matiere_premiere_id, proportion in dosages:
        cursor.execute("SELECT quantite FROM matieres_premieres WHERE id = ?", (matiere_premiere_id,))
        quantite_disponible = cursor.fetchone()[0]
        quantite_necessaire = quantite * proportion

        if quantite_disponible < quantite_necessaire:
            # Annuler toutes les modifications précédentes
            for ancien_id, ancienne_quantite in modifications:
                cursor.execute("UPDATE matieres_premieres SET quantite = quantite + ? WHERE id = ?", 
                            (ancienne_quantite, ancien_id))
            raise ValueError(f"Quantité insuffisante pour la matière première ID {matiere_premiere_id}. Les modifications ont été annulées.")

        # Effectuer la réduction et enregistrer la modification
        cursor.execute("UPDATE matieres_premieres SET quantite = quantite - ? WHERE id = ?", 
                    (quantite_necessaire, matiere_premiere_id))
        modifications.append((matiere_premiere_id, quantite_necessaire))

    # Ajouter la quantité produite
    cursor.execute("UPDATE matieres_produites SET quantite = quantite + ?, prix_unitaire = ?, fodec = ? WHERE reference = ?", 
                (quantite, prix, fodec_value, matiere_produite_reference))
    # Ajouter l'entrée dans l'historique
    cursor.execute("INSERT INTO historique_matiere_produite (nom, quantite, prix_unitaire) VALUES (?, ?, ?)", 
                (nom, quantite, prix))
    
    conn.commit()
    conn.close()


def creer_bon_livraison(articles, client, client_info, bl_value):
    """
    Crée un bon de livraison pour plusieurs matières produites.
    Annule les modifications en cas d'erreur.
    :param articles: Liste de dictionnaires {"matiere_produite_nom": str, "quantite": float}
    :param client: Nom du client
    :return: Contenu de la facture générée
    """
    conn = get_connection()
    cursor = conn.cursor()

    try:
        # Démarrer une transaction
        conn.execute("BEGIN TRANSACTION")

        date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Créer un bon de livraison (sans les détails pour l'instant)
        cursor.execute("INSERT INTO bons_livraison (date, client) VALUES (?, ?)", (date, client))
        bon_id = cursor.lastrowid

        total_general = 0.0  # Pour calculer le montant total de la facture
        details_facture = []  # Stocker les lignes de facture
        articles_historique = []  # Stocker les détails des articles pour l'historique
        produits = [] # Stocker les matières produites à livrais

        for article in articles:
            matiere_produite_nom = article["matiere_produite_nom"]
            quantite = article["quantite"]

            # Vérifier la disponibilité de la matière produite
            cursor.execute("SELECT id, quantite, prix_unitaire, fodec FROM matieres_produites WHERE nom = ?", (matiere_produite_nom,))
            result = cursor.fetchone()
            if not result:
                raise ValueError(f"Matière produite '{matiere_produite_nom}' non trouvée")
            
            matiere_produite_id, quantite_disponible, prix_unitaire, fodec = result

            if quantite_disponible < quantite:
                raise ValueError(f"Quantité insuffisante pour '{matiere_produite_nom}'")
                
            # Réduire la quantité de la matière produite
            cursor.execute(
                "UPDATE matieres_produites SET quantite = quantite - ? WHERE id = ?", 
                (quantite, matiere_produite_id)
            )

            # Ajouter les détails du bon de livraison
            cursor.execute(
                "INSERT INTO details_bon_livraison (bon_livraison_id, matiere_produite_id, quantite) VALUES (?, ?, ?)", 
                (bon_id, matiere_produite_id, quantite)
            )

            # Calculer le total pour cet article
            total_article = prix_unitaire * quantite
            total_general += total_article

            # Ajouter les détails à la facture
            details_facture.append(f"{matiere_produite_nom} (Quantité: {quantite}, Prix/Unité: {prix_unitaire}, Total: {total_article:.2f})")
            produits.append({'nom': matiere_produite_nom, 'quantite': quantite, 'prix': prix_unitaire, 'fodec': fodec})

            # Ajouter les articles à l'historique
            articles_historique.append({
                "nom": matiere_produite_nom,
                "quantite": quantite,
                "prix_unitaire": prix_unitaire,
                "total": total_article
            })

        # Ajouter une entrée dans l'historique
        cursor.execute(
            "INSERT INTO historique_bon_livraison (bon_id, client, date_creation, total, articles) VALUES (?, ?, ?, ?, ?)", 
            (bon_id, client, date, total_general, json.dumps(articles_historique))
        )

        # Valider la transaction si tout est correct
        conn.commit()

        # Générer et retourner la facture
         # Chemin du dossier de l'utilisateur
        user_dir = Path.home() / "GestionStockApp"
        F_model = user_dir / 'FACTURE-Copy.xlsx'
        if bl_value == '1' :
            cursor.execute("SELECT MAX(num_bl) FROM bons_livraison")
            num_bl = cursor.fetchone()[0] or 0  # Si aucun résultat, initialisez à 1
            print('bl =',bl_value,' num_bl = ',num_bl , ' bon_id =' , bon_id)
            cursor.execute("UPDATE bons_livraison SET num_bl = ? WHERE id = ?", (num_bl+1, bon_id))
            generer_facture2(num_bl+1, client_info, produits, F_model, date, bl_value)
            return generer_facture(num_bl+1, client, date, details_facture, total_general)
        else :
            cursor.execute("SELECT MAX(num_facture) FROM bons_livraison")
            num_facture = cursor.fetchone()[0] or 0  # Si aucun résultat, initialisez à 1
            cursor.execute("UPDATE bons_livraison SET num_facture = ? WHERE id = ?", (num_facture+1, bon_id))
            generer_facture2(num_facture+1, client_info, produits, F_model, date, bl_value)
            return generer_facture(num_facture+1, client, date, details_facture, total_general)

    except Exception as e:
        # Annuler toutes les modifications si une erreur survient
        conn.rollback()
        raise e

    finally:
        # Fermer la connexion
        conn.commit()
        conn.close()


def generer_facture(bon_id, client, date, details, total_general):
    """
    Génère une facture consolidée pour un bon de livraison.
    :param bon_id: ID du bon de livraison
    :param client: Nom du client
    :param date: Date de création
    :param details: Liste des détails (articles) de la facture
    :param total_general: Total général de la facture
    :return: Contenu de la facture générée
    """
    # Construire le contenu de la facture
    facture = f"""
    BON DE LIVRAISON N° {bon_id}
    ---------------------------
    Client : {client}
    Date : {date}

    Articles Livrés :
    -----------------
    """
    for detail in details:
        facture += f"- {detail}\n"

    facture += f"""
    ---------------------------
    Total Général : {total_general:.2f}

    Merci pour votre collaboration.
    """

    # Sauvegarder la facture dans un fichier texte
    # with open(f"facture_{bon_id}.txt", "w", encoding="utf-8") as file:
    #     file.write(facture)

    return facture

# autre methode
def generer_facture2(bon_id, client_info, produits, fichier_modele, date, bl_value):
    """
    Génère une facture Excel en remplissant les informations du client et les produits finis.

    :param client_info: Dictionnaire contenant 'nom', 'adresse', 'phone', 'reference'.
    :param produits: Liste de dictionnaires {'nom': str, 'quantite': int, 'prix': float, 'fodec': str }.
    :param fichier_modele: Chemin du fichier modèle Excel.
    """
    # Charger le fichier modèle avec openpyxl
    wb = openpyxl.load_workbook(fichier_modele)
    ws = wb.active

    # Insérer les informations du client (E3 à H6)
    ws["E3"] = client_info["nom"]
    ws["E4"] = client_info["adresse"]
    ws["E5"] = client_info["phone"]
    ws["E6"] = 'MF ' + client_info["reference"]

    # Insérer la date à B7
    date_obj = datetime.strptime(date, "%Y-%m-%d %H:%M:%S")
    formatted_date = date_obj.strftime("%d/%m/%Y")
    ws["B7"] = "Date : " + str(formatted_date)

    # Insérer l'ID à B6
    if bl_value == '1' :
        ws["B6"] = 'BL/Facture N° :' + str(bon_id) + '/' + str(date_obj.strftime("%y"))
    else :
        ws["B6"] = 'Facture N° :' + str(bon_id) + '/' + str(date_obj.strftime("%y"))

    # Insérer les produits finis à partir de B10
    ligne_depart = 10
    total_general = 0

    for i, produit in enumerate(produits):
        ligne = ligne_depart + i
        ws[f"B{ligne}"] = produit["nom"]
        ws[f"D{ligne}"] = produit["quantite"]
        ws[f"E{ligne}"] = produit["prix"]
        ws[f"G{ligne}"] = '1%' if produit["fodec"]=='1' else ''
        total_produit = produit["quantite"] * produit["prix"]
        ws[f"F{ligne}"] = total_produit
        ws[f"H{ligne}"] = '19%'
        total_general += total_produit

    # Insérer le total en F38 à H38
    #ws["F38"] = total_general


    dossier_facture = "./factures"  # Chemin relatif pour le dossier facture
    os.makedirs(dossier_facture, exist_ok=True)  # Crée le dossier s'il n'existe pas
    if bl_value == '1' :
        fichier_sortie = os.path.join(
            dossier_facture, f"BL_FACTURE_{str(bon_id)} {str(date_obj.strftime('%d-%m-%Y'))}.xlsx"
        )
    else :
        fichier_sortie = os.path.join(
            dossier_facture, f"FACTURE_{str(bon_id)} {str(date_obj.strftime('%d-%m-%Y'))}.xlsx"
        )
    print(f"Facture sauvegardée dans : {fichier_sortie}")

    # Sauvegarder temporairement pour xlwings
    # fichier_sortie = f"FACTURE_{str(bon_id)} {str(date_obj.strftime('%d-%m-%Y'))}.xlsx"
    wb.save(fichier_sortie)
    wb.close()

    # Ouvrir le fichier avec xlwings pour recalculer les formules
    app = xw.App(visible=False)
    wb_xlwings = app.books.open(fichier_sortie)
    ws_xlwings = wb_xlwings.sheets[0]

    # Forcer le recalcul des formules dans Excel
    wb_xlwings.app.calculate()

    # Lire la valeur exacte de la cellule F42
    totale = round(ws_xlwings["F42"].value, 3)
    print("Valeur calculée dans F42 :", totale)

    # Convertir le total en lettres et l’insérer en B38
    partie_entiere, partie_decimale = str(totale).split(".")
    texte_entiere = num2words(int(partie_entiere), lang="fr")
    partie_decimale = partie_decimale.ljust(3, '0')[:3]
    texte_decimale = num2words(int(partie_decimale), lang="fr")
    total_en_lettres = f"{texte_entiere} Dinars, {texte_decimale} Millimes."
    ws_xlwings["B38"].value = 'Arrêtée la présente facture à la somme de  : ' + total_en_lettres

    # Sauvegarder et fermer
    wb_xlwings.save(fichier_sortie)
    wb_xlwings.close()
    app.quit()

    return fichier_sortie